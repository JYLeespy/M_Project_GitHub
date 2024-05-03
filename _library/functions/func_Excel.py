#GitHub Update 2024-05-03
import os
import xlwings as xw
import openpyxl
from openpyxl.styles import Alignment
from openpyxl import load_workbook

class Excel:
    def __init__(self, filepath):
        #WorkBook 생성
        self.wb = openpyxl.Workbook()
        self.wb.save(filepath)
        print(f'[Excel] Workbook 생성 {filepath}')
        # self.wb = xw.Book(filepath)

    def create_sheet(self, sheetname):
        self.wb.create_sheet(sheetname)
    def insert_row(self, sheetname, row):
        self.ws = self.wb[sheetname]
        self.ws.append(row)

    def cell_merge(self, range):
        self.ws.merge_cells(range)
        self.ws.alignment = Alignment(horizontal= 'center', vertical='center')

    def save(self, savefilename):
        self.wb.save(savefilename)




class ExcelWings:
    def __init__(self, filepath=""):
        self.wb = xw.Book(filepath)

    def sheetnames(self):
        return self.wb.sheet_names

    def sheetadd(self, sheetname, zoom = 100, showGrid=True):
        print(f'sheet add : {sheetname}')
        if len(sheetname) > 31:
            print(f'sheet 이름이 너무 깁니다. 최대 31 자릿 수 까지 가능합니다. {sheetname}')
            sheetname = sheetname[0:31]
        try:
            self.wb.sheets.add(sheetname)
            app = xw.apps.active
            app.api.ActiveWindow.DisplayGridlines = showGrid
            app.api.ActiveWindow.Zoom = zoom
            return True
        except ValueError:
            print(f'[sheetadd Error] 현재 Workbook에 이미 "{sheetname}"가 있습니다.')
            return False

    def coumns_width(self, sheetname, range, width):
        if len(sheetname) > 31:
            sheetname = sheetname[0:31]
        ws = self.wb.sheets[sheetname]
        ws.range(range).column_width = width
        # 자동 맞추기
        # ws.Columns.AutoFit()
    def insert(self, sheetname, insert_range, listdata):
        if len(sheetname) > 31:
            sheetname = sheetname[0:31]
        ws = self.wb.sheets[sheetname]
        ws.range(insert_range).value = listdata
    def inputdata(self, sheetname, cell, data):
        if len(sheetname) > 31:
            sheetname = sheetname[0:31]
        ws = self.wb.sheets[sheetname]
        ws.range(cell).value = data

    def readdata(self, sheetname, cell):
        if len(sheetname) > 31:
            sheetname = sheetname[0:31]
        ws = self.wb.sheets[sheetname]
        return ws.range(cell).value

    def formula(self, sheetname, cell, fomula_data):
        if len(sheetname) > 31:
            sheetname = sheetname[0:31]
        ws = self.wb.sheets[sheetname]
        ws.range(cell).formula = fomula_data

    def border(self, sheetname, range, border='all', weight=1, color='#ffffff'):
        if len(sheetname) > 31:
            sheetname = sheetname[0:31]
        ws = self.wb.sheets[sheetname]
        ws_range = ws.range(range)
        # ws_range.api.Borders(border).LineStyle = linestyle

        boder_style = {'left':[1],
                       'right':[2],
                       'top':[3],
                       'bottom':[4],
                       'diagonal-lefttop':[5],
                       'diagonal-righttop':[6],
                       'all':[1,2,3,4],
                       'x':[5,6]}

        for i in boder_style[border]:
            ws_range.api.Borders(i).Weight = weight

        # ws_range.api.Borders(border).Color = color
        # xlDiagonalDown = 5  # from enum XlBordersIndex
        # xlDiagonalUp = 6  # from enum XlBordersIndex
        # xlEdgeBottom = 9  # from enum XlBordersIndex
        # xlEdgeLeft = 7  # from enum XlBordersIndex
        # xlEdgeRight = 10  # from enum XlBordersIndex
        # xlEdgeTop = 8  # from enum XlBordersIndex
        # xlInsideHorizontal = 12  # from enum XlBordersIndex
        # xlInsideVertical = 11  # from enum XlBordersIndex

    def chart(self, sheetname, range, position, chart_type='bar', chart_height=200, chart_width=1000):
        if len(sheetname) > 31:
            sheetname = sheetname[0:31]
        ws = self.wb.sheets[sheetname]
        chart = ws.charts.add()
        chart.set_source_data(ws.range(range).expand())

        chart.height = chart_height
        chart.width = chart_width

        #chart.chart_type = 'line'

        chart.top = ws.range(position).top
        chart.left = ws.range(position).left

    def image_insert(self, img_file, sheetname,  col, row, resize):
        if len(sheetname) > 31:
            sheetname = sheetname[0:31]
        try:
            ws = self.wb.sheets[sheetname]
        except:
            self.sheetadd(sheetname)
            ws = self.wb.sheets[sheetname]
        cell = f'{col}{row}'
        # print(cell, resize)
        if resize[1] == '0':
            if resize[0] == "h":
                ws.pictures.add(image=os.path.abspath(img_file), left=ws.range(cell).left, top=ws.range(cell).top, height=ws[cell].api.RowHeight)
            else:
                ws.pictures.add(image=os.path.abspath(img_file), left=ws.range(cell).left, top=ws.range(cell).top, width=ws[cell].api.ColumnWidth*6.37142857142857)
        else:
            length = float(resize[1])
            if resize[0] == 'h':
                ws[cell].api.RowHeight = length
                ws.pictures.add(os.path.abspath(img_file), left=ws.range(cell).left, top=ws.range(cell).top, height=length)
            else:
                ws[cell].api.ColumnWidth = length
                ws.pictures.add(os.path.abspath(img_file), left=ws.range(cell).left, top=ws.range(cell).top, width=length*6.37142857142857)
    def convert_cols(self, col_number):
        column_trans = ['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z']
        if col_number//26 == 0:
            return column_trans[col_number % 26]
        elif col_number//26 >= 1 and col_number//26 <= 25:
            return f'{column_trans[((col_number // 26))-1]}{column_trans[col_number % 26]}'
    def reverse_cols(self, col_str) -> int:
        column_trans = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S',
                        'T', 'U', 'V', 'W', 'X', 'Y', 'Z']
        col_num = 0
        for multifplier in range(len(col_str)):
            # digit : 자릿수
            digit = len(col_str)-multifplier-1
            col_num += (26**multifplier)*(column_trans.index(col_str[digit])+1)
        return col_num

    def save(self, saveas, password=None):
        self.wb.save(saveas, password)
    def close(self):
        self.wb.close()